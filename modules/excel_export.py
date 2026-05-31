"""Excel export — professional multi-sheet workbook."""
from __future__ import annotations
from io import BytesIO
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

NAVY = "FF2B0000"
BURGUNDY = "FF450909"
BLUE = "FF7ED6FF"
RED = "FFFF2B4D"
WHITE = "FFFFFFFF"
GREY = "FFD9D9D9"

thin = Side(border_style="thin", color="FF555555")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
        cell.fill = PatternFill("solid", fgColor=BURGUNDY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _autosize(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 12), 32)


def _write_df(ws, df: pd.DataFrame, start_row: int = 1,
              money_cols: list[str] | None = None,
              pct_cols: list[str] | None = None) -> int:
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="No data").font = Font(italic=True, color=GREY)
        return start_row + 1
    money_cols = money_cols or []
    pct_cols = pct_cols or []

    # headers
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    _style_header(ws, start_row, len(df.columns))

    # body
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            cell = ws.cell(row=i, column=j, value=(None if pd.isna(v) else v))
            cell.font = Font(color=WHITE, name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.border = BORDER
            if col in money_cols:
                cell.number_format = '_-* #,##0_-;[Red]-#,##0;_-* "-"_-'
                cell.alignment = Alignment(horizontal="right")
            elif col in pct_cols:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            elif col == "Year":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")
    ws.freeze_panes = ws.cell(row=start_row + 1, column=2)
    return start_row + len(df) + 1


def build_workbook(setup: dict, forecast_df: pd.DataFrame,
                   pnl: pd.DataFrame, bs: pd.DataFrame, cf: pd.DataFrame,
                   roi_df: pd.DataFrame, summary: dict,
                   logo_path: str | Path | None = None) -> bytes:
    wb = Workbook()

    # ---------- Executive Summary ----------
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    if logo_path and Path(logo_path).exists():
        try:
            img = XLImage(str(logo_path))
            img.width = 110
            img.height = 110
            ws.add_image(img, "A1")
        except Exception:
            pass

    ws["C2"] = "FIN APPS — Investment Analysis"
    ws["C2"].font = Font(bold=True, size=20, color=BLUE, name="Calibri")
    ws["C3"] = "Financial Insights. Better Decisions."
    ws["C3"].font = Font(italic=True, color=GREY, size=11)
    ws["C4"] = f"Project: {setup.get('project_name','—')}    Currency: {setup.get('currency','USD')}"
    ws["C4"].font = Font(color=WHITE, size=11)

    kpi_rows = [
        ("Total Investment", summary.get("total_investment", 0), "money"),
        ("IRR", summary.get("irr"), "pct"),
        ("NPV", summary.get("npv", 0), "money"),
        ("Payback (yrs)", summary.get("payback"), "num"),
        ("Discounted Payback (yrs)", summary.get("discounted_payback"), "num"),
        ("Average EBITDA", summary.get("ebitda_avg", 0), "money"),
        ("Average ROI", summary.get("roi_avg", 0), "pct"),
        ("Peak Working Capital", summary.get("peak_wc", 0), "money"),
        ("Cumulative Cash Flow", summary.get("cum_cashflow", 0), "money"),
        ("Discount Rate (Hurdle)", summary.get("hurdle", 0), "pct"),
    ]
    start = 7
    ws.cell(row=start, column=2, value="Key Metric").font = Font(bold=True, color=WHITE)
    ws.cell(row=start, column=3, value="Value").font = Font(bold=True, color=WHITE)
    for c in (2, 3):
        ws.cell(row=start, column=c).fill = PatternFill("solid", fgColor=BURGUNDY)
        ws.cell(row=start, column=c).border = BORDER
        ws.cell(row=start, column=c).alignment = Alignment(horizontal="center")
    for i, (lbl, val, kind) in enumerate(kpi_rows, start=start + 1):
        ws.cell(row=i, column=2, value=lbl).font = Font(color=GREY)
        c = ws.cell(row=i, column=3, value=val if val is not None else "—")
        c.font = Font(bold=True, color=BLUE if (isinstance(val,(int,float)) and val and val >= 0) else RED, size=12)
        if kind == "money":
            c.number_format = '_-* #,##0_-;[Red]-#,##0;_-* "-"_-'
        elif kind == "pct":
            c.number_format = "0.0%"
        else:
            c.number_format = "0.00"
        for col in (2, 3):
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(row=i, column=col).border = BORDER
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 24

    # ---------- Assumptions ----------
    ws2 = wb.create_sheet("Assumptions")
    ws2.sheet_view.showGridLines = False
    assumptions = pd.DataFrame([
        {"Parameter": "Project Name", "Value": setup.get("project_name", "")},
        {"Parameter": "Currency", "Value": setup.get("currency", "USD")},
        {"Parameter": "Start Year", "Value": setup.get("start_year", "")},
        {"Parameter": "Number of Years", "Value": setup.get("num_years", "")},
        {"Parameter": "Initial Capex", "Value": setup.get("initial_capex", 0)},
        {"Parameter": "Depreciable Capex", "Value": setup.get("depreciable_capex", 0)},
        {"Parameter": "Useful Life (years)", "Value": setup.get("useful_life", 0)},
        {"Parameter": "Residual Value", "Value": setup.get("residual_value", 0)},
        {"Parameter": "Tax Rate", "Value": setup.get("tax_rate", 0)},
        {"Parameter": "Discount Rate / WACC", "Value": setup.get("discount_rate", 0)},
        {"Parameter": "Receivable Days", "Value": setup.get("receivable_days", 0)},
        {"Parameter": "Inventory Days", "Value": setup.get("inventory_days", 0)},
        {"Parameter": "Payable Days", "Value": setup.get("payable_days", 0)},
        {"Parameter": "Recover Working Capital", "Value": setup.get("recover_working_capital", True)},
    ])
    _write_df(ws2, assumptions)
    _autosize(ws2)

    # ---------- Forecast Input ----------
    ws3 = wb.create_sheet("Forecast Input")
    ws3.sheet_view.showGridLines = False
    _write_df(ws3, forecast_df,
              money_cols=["Sales", "Material Value", "Direct Labour Value",
                          "MOH Value", "SG&A Value"],
              pct_cols=["Material %", "Direct Labour %", "MOH %", "SG&A %"])
    _autosize(ws3)

    # ---------- P&L ----------
    ws4 = wb.create_sheet("P&L")
    ws4.sheet_view.showGridLines = False
    _write_df(ws4, pnl,
              money_cols=["Sales","Material Cost","Gross Profit","Direct Labour",
                          "Manufacturing Overhead","SG&A","EBITDA","Depreciation",
                          "EBIT","Taxes","Net Income"],
              pct_cols=["EBITDA Margin","Net Margin"])
    _autosize(ws4)

    # ---------- Balance Sheet ----------
    ws5 = wb.create_sheet("Balance Sheet")
    ws5.sheet_view.showGridLines = False
    _write_df(ws5, bs,
              money_cols=["Gross Fixed Assets","Accumulated Depreciation","Net Fixed Assets",
                          "Receivables","Inventory","Payables","Net Working Capital","Net Investment"])
    _autosize(ws5)

    # ---------- Cash Flow ----------
    ws6 = wb.create_sheet("Cash Flow")
    ws6.sheet_view.showGridLines = False
    _write_df(ws6, cf,
              money_cols=["Net Income","Depreciation","Change in WC","Capex",
                          "Residual Value","Cash Flow","Cumulative Cash Flow"])
    _autosize(ws6)

    # ---------- IRR & ROI ----------
    ws7 = wb.create_sheet("IRR & ROI")
    ws7.sheet_view.showGridLines = False
    metrics = pd.DataFrame([
        {"Metric": "IRR", "Value": summary.get("irr")},
        {"Metric": "NPV", "Value": summary.get("npv")},
        {"Metric": "Payback (yrs)", "Value": summary.get("payback")},
        {"Metric": "Discounted Payback (yrs)", "Value": summary.get("discounted_payback")},
        {"Metric": "Average ROI", "Value": summary.get("roi_avg")},
        {"Metric": "3-Year ROI", "Value": summary.get("roi_3y")},
        {"Metric": "5-Year ROI", "Value": summary.get("roi_5y")},
    ])
    next_row = _write_df(ws7, metrics)
    next_row += 1
    ws7.cell(row=next_row, column=1, value="ROI by Year").font = Font(bold=True, color=BLUE, size=13)
    _write_df(ws7, roi_df, start_row=next_row + 1,
              money_cols=["EBIT","Average Investment"], pct_cols=["ROI"])
    _autosize(ws7)

    # ---------- Charts placeholder ----------
    ws8 = wb.create_sheet("Charts")
    ws8.sheet_view.showGridLines = False
    ws8["B2"] = "Interactive charts are available in the FIN APPS web app."
    ws8["B2"].font = Font(color=GREY, italic=True, size=12)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
